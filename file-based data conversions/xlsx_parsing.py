# activate venv using:
# source venv/bin/activate

# to stop use:
# deactivate

# document installed dependencies using:
# pip freeze > requirements.txt

# to install dependencies in another venv, use:
# pip install -r requirements.txt

from openpyxl import load_workbook, Workbook
import csv

source_workbook: Workbook = load_workbook(filename='fredgraph.xlsx')
print(source_workbook.sheetnames)

for sheet_num, sheet_name in enumerate(source_workbook.sheetnames):
    current_sheet = source_workbook[sheet_name]
    print(sheet_name)
    print(sheet_num)
    
    output_file = open("xlsx_"+sheet_name+".csv", "w")
    output_writer = csv.writer(output_file)

    for row in current_sheet.iter_rows():
        row_cells = []
        for cell in row:
            print(cell, cell.value)
            row_cells.append(cell.value)
        output_writer.writerow(row_cells)
    output_file.close()